from typing import List, Dict, Any

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from logger import logger
from utils.common_utils import get_color_code_by_number
from utils.datetime_utils import *
from utils.request_utils import get_request_resource_as_json


def merge_dataframes(
        left_df: pd.DataFrame,
        right_df: pd.DataFrame,
        merge_type: str,
        merge_columns: str | List[str],
        suffixes: tuple
) -> pd.DataFrame:
    """
    Merges two dataframes by given columns.

    Parameters
    ----------
    left_df : pd.DataFrame
        First DataFrame to merge.
    right_df : pd.DataFrame
        Second DataFrame to merge with.
    merge_type : str
        How to merge the two DataFrames.
    merge_columns: str
        Columns names to join on.
    suffixes : tuple
        String indicating the suffix to add to overlapping column names in left and right respectively.

    Returns
    -------
    new_df : pd.DataFrame
        The merged DataFrame.
    """
    new_df = left_df.merge(right_df, how=merge_type, on=merge_columns, suffixes=suffixes)

    logger.info("DataFrames merged successfully!")

    return new_df


def filter_rows_with_null_values_from_df(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """
    Removes rows from DataFrame where value from given column is Null.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to filter.
    column : str
        Column name to use for filtering.

    Returns
    -------
    new_df : pd.DataFrame
        The filtered DataFrame.
    """
    new_df = df[df[column].notnull()]

    logger.info(f"Filtered Null values from {column} column from DataFrame.")

    return new_df


def get_common_columns_from_dfs(first_df: pd.DataFrame, second_df: pd.DataFrame) -> list:
    """
    Finds the common column names from two DataFrames.

    Parameters
    ----------
    first_df : pd.DataFrame
        First DataFrame to compare.
    second_df : pd.DataFrame
        Second DataFrame to compare.

    Returns
    -------
    common_columns : list
        List with the common column names.
    """
    common_columns = first_df.columns.intersection(second_df.columns).tolist()

    return common_columns


def replace_null_values_in_df(df: pd.DataFrame, common_columns: List[str], suffix: str) -> pd.DataFrame:
    """
    Compares the generated columns in DataFrame after merge
    and check if one is Null, then takes the value from the other.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to use for replacing the values.
    common_columns : List[str]
        List with column names to compare.
    suffix : str
        Suffix extension of the duplicate columns

    Returns
    -------
    df : pd.DataFrame
        DataFrame with replaced values.
    """
    for column in common_columns:
        df[column] = df[column].where(df[column].notnull(), df[column + suffix])

    logger.info("Replaced Null values in DataFrame.")

    return df


def drop_suffix_columns(df: pd.DataFrame, columns: List[str], suffix: str) -> pd.DataFrame:
    """
    Drops redundant columns from DataFrame generated after merge.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to use for dropping the columns.
    columns : List[str]
        List with column names to drop.
    suffix : str
        Suffix extension of the redundant columns

    Returns
    -------
    df : pd.DataFrame
        DataFrame without redundant columns.
    """
    new_df = df.drop([f"{c}{suffix}" for c in columns], axis=1)

    logger.info("Dropped redundant columns from DataFrame.")

    return new_df


def drop_mismatch_columns(df: pd.DataFrame, columns: List[str]):
    """
    Compares and drop columns from DataFrame
    that are not in the given list.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to use for dropping the columns.
    columns : List[str]
        List with column names to compare.

    Returns
    -------
    df : pd.DataFrame
        DataFrame without the columns that are not in the list.
    """
    new_df = df.drop(df.columns.difference(columns, sort=False), axis=1)

    logger.info("Dropped mismatch columns from DataFrame.")

    return new_df


def sort_dataframe(df: pd.DataFrame, sort_columns: str | List[str], ascending: bool | List[bool]) -> pd.DataFrame:
    """
    Sorts DataFrame by given columns in given order.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to sort.
    sort_columns : List[str]
        List with column names to use for sorting.
    ascending: bool
        How to sort the DataFrame.

    Returns
    -------
    df : pd.DataFrame
        Sorted Dataframe.
    """
    new_df = df.sort_values(by=sort_columns, ascending=ascending)

    logger.info(f"DataFrame sorted in {'ascending' if ascending else 'descending'} order by {sort_columns} column.")

    return new_df


def add_new_column_to_dataframe(df: pd.DataFrame, column_name: str, value: Any = None):
    """
    Adds new column in DataFrame with given value.

    Parameters
    ----------
    df: pd.DataFrame
        DataFrame to use for adding the new column.
    column_name: str
        Name of the new column.
    value: Any
        Value to insert in the new column, None by default.

    Returns
    -------
    df: pd.DataFrame
        Dataframe with the new column.
    """
    df[column_name] = value

    logger.info(f"Added {column_name} column in DataFrame.")

    return df


def write_data_from_pandas_dataframe_to_worksheet(
        df: pd.DataFrame,
        ws: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    """
    Writes data from pandas Dataframe to openpyxl Worksheet.

    Parameters
    ----------
    df: pd.DataFrame
        DataFrame to use for getting the data.
    ws: openpyxl.worksheet.worksheet.Worksheet
        Worksheet to use for writing the data.

    Returns
    -------
    None
    """
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[1]:
        cell.style = 'Pandas'

    logger.info("Data written from DataFrame to Worksheet successfully!")


def get_color_codes(
        label_ids: pd.Series,
        url: str,
        headers=Dict[str, str]
) -> List[str]:
    """
    Extracts color codes from API via request using values from array.

    Parameters
    ----------
    label_ids: pd.Series
        Array containing the values.
    url: str
        URL of the color codes.
    headers: Dict[str, str]
        Headers to send with the request.

    Returns
    -------
    color_codes: List[str]
        List with all the extracted color codes
    """
    color_codes = []

    for label_id in label_ids:
        if label_id is not None:
            current_url = f"{url}{label_id}"
            resource = get_request_resource_as_json(current_url, headers)
            try:
                current_code = resource[0]["colorCode"]
                color_codes.append(current_code)
            except (IndexError, KeyError):
                pass

    return color_codes


def add_font_color_to_worksheet_cells(
        label_ids: pd.Series,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        url: str,
        headers: Dict[str, str]
) -> None:
    """
    Adds font color to Worksheet's cells text by
    extracting color code from API using values from array.

    Parameters
    ----------
    label_ids: pd.Series
        Array containing the values.
    ws: openpyxl.worksheet.worksheet.Worksheet
        Worksheet with the data.
    url: str
        URL of the color codes.
    headers: Dict[str, str]
        Headers to send with the request.

    Returns
    -------
    None
    """
    color_codes = get_color_codes(label_ids, url, headers)
    if not color_codes == []:
        first_code = color_codes[0].lstrip("#")
        rows_range = ws.max_row
        columns_range = ws.max_column
        for row_index in range(1, rows_range + 1):
            for column_index in range(1, columns_range + 1):
                ws.cell(row_index, column_index).font = openpyxl.styles.Font(color=first_code)

        logger.info("Added font color to Worksheet.")


def get_column_index_in_worksheet(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        required_column,
        rows_range,
        columns_range
) -> int:
    """
    Finds the column index of column in Worksheet.

    Parameters
    ----------
    ws: openpyxl.worksheet.worksheet.Worksheet
        Worksheet with the data.
    required_column: str
        Name of the required column
    rows_range:
        Worksheet's rows count
    columns_range:
        Worksheet's columns count

    Returns
    -------
    column_index: int
        The index of the required column
    """
    for row_index in range(1, rows_range + 1):
        for column_index in range(1, columns_range + 1):
            if ws.cell(row_index, column_index).value == required_column:
                return column_index


def add_background_color_to_worksheet_cells(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        required_column: str,
        today: datetime
) -> None:
    """
    Adds background color to Worksheet's cells
    depending on months count between two dates.

    Parameters
    ----------
    ws: openpyxl.worksheet.worksheet.Worksheet
        Worksheet with the data.
    required_column: str
        Column to use for start date
    today: datetime
        End date

    Returns
    -------
    None
    """
    rows_range = ws.max_row
    columns_range = ws.max_column

    hu_column_index = get_column_index_in_worksheet(ws, required_column, rows_range, columns_range)

    for row_index in range(1, rows_range + 1):
        if hu_column_index is not None and row_index > 1:
            date_str = ws.cell(row_index, hu_column_index).value
            date = string_to_date(date_str)
            months_diff = get_months_diff_between_dates(date, today)
            color_code = get_color_code_by_number(months_diff)
            for column_index in range(1, columns_range + 1):
                cell_header = ws.cell(row_index, column_index)
                cell_header.fill = PatternFill(start_color=color_code, fill_type="solid")

    logger.info("Added background color to Worksheet.")
